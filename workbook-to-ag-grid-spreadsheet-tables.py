import pandas as pd
import openpyxl as op
import numpy as np
import re
import collections

#FOR COLOR FINDING
#Credit to Andrew Thornton @ https://bitbucket.org/openpyxl/openpyxl/issues/987/add-utility-functions-for-colors-to-help
from colorsys import rgb_to_hls, hls_to_rgb

RGBMAX = 0xff  # Corresponds to 255
HLSMAX = 240  # MS excel's tint function expects that HLS is base 240. see:
# https://social.msdn.microsoft.com/Forums/en-US/e9d8c136-6d62-4098-9b1b-dac786149f43/excel-color-tint-algorithm-incorrect?forum=os_binaryfile#d3c2ac95-52e0-476b-86f1-e2a697f24969

def rgb_to_ms_hls(red, green=None, blue=None):
  """Converts rgb values in range (0,1) or a hex string of the form '[#aa]rrggbb' to HLSMAX based HLS, (alpha values are ignored)"""
  if green is None:
    if isinstance(red, str):
      if len(red) > 6:
        red = red[-6:]  # Ignore preceding '#' and alpha values
      blue = int(red[4:], 16) / RGBMAX
      green = int(red[2:4], 16) / RGBMAX
      red = int(red[0:2], 16) / RGBMAX
    else:
      red, green, blue = red
  h, l, s = rgb_to_hls(red, green, blue)
  return (int(round(h * HLSMAX)), int(round(l * HLSMAX)), int(round(s * HLSMAX)))

def ms_hls_to_rgb(hue, lightness=None, saturation=None):
  """Converts HLSMAX based HLS values to rgb values in the range (0,1)"""
  if lightness is None:
    hue, lightness, saturation = hue
  return hls_to_rgb(hue / HLSMAX, lightness / HLSMAX, saturation / HLSMAX)

def rgb_to_hex(red, green=None, blue=None):
  """Converts (0,1) based RGB values to a hex string 'rrggbb'"""
  if green is None:
    red, green, blue = red
  return ('%02x%02x%02x' % (int(round(red * RGBMAX)), int(round(green * RGBMAX)), int(round(blue * RGBMAX)))).upper()

def get_theme_colors(wb):
  """Gets theme colors from the workbook"""
  # see: https://groups.google.com/forum/#!topic/openpyxl-users/I0k3TfqNLrc
  from openpyxl.xml.functions import QName, fromstring
  xlmns = 'http://schemas.openxmlformats.org/drawingml/2006/main'
  root = fromstring(wb.loaded_theme)
  themeEl = root.find(QName(xlmns, 'themeElements').text)
  colorSchemes = themeEl.findall(QName(xlmns, 'clrScheme').text)
  firstColorScheme = colorSchemes[0]

  colors = []

  for c in ['lt1', 'dk1', 'lt2', 'dk2', 'accent1', 'accent2', 'accent3', 'accent4', 'accent5', 'accent6']:
    accent = firstColorScheme.find(QName(xlmns, c).text)

    if 'window' in list(accent)[0].attrib['val']:
      colors.append(list(accent)[0].attrib['lastClr'])
    else:
      colors.append(list(accent)[0].attrib['val'])

  return colors

def tint_luminance(tint, lum):
  """Tints a HLSMAX based luminance"""
  # See: http://ciintelligence.blogspot.co.uk/2012/02/converting-excel-theme-color-and-tint.html
  if tint < 0:
    return int(round(lum * (1.0 + tint)))
  else:
    return int(round(lum * (1.0 - tint) + (HLSMAX - HLSMAX * (1.0 - tint))))

def theme_and_tint_to_rgb(wb, theme, tint):
  """Given a workbook, a theme number and a tint return a hex based rgb"""
  rgb = get_theme_colors(wb)[theme]
  h, l, s = rgb_to_ms_hls(rgb)
  return rgb_to_hex(ms_hls_to_rgb(h, tint_luminance(tint, l), s))

def getBgColor(wb, obj):
  
  scolor = obj.fgColor
  
  color = ''

  if obj.patternType != None:
    if scolor.type == 'theme':
      color = theme_and_tint_to_rgb(wb, scolor.theme, scolor.tint)
    elif scolor.type == 'rgb':
      color = scolor.rgb[2:]
    else:
      color = ''
      #print('bg')
      #print(obj)

  return color

def getFtColor(wb, obj):

  scolor = obj.color

  color = ''

  if scolor != None:
    if scolor.type == 'theme':
      color = theme_and_tint_to_rgb(wb, scolor.theme, scolor.tint)
    elif scolor.type == 'rgb':
      color = scolor.rgb[2:]
    else:
      color = ''
      #print('ft')
      #print(obj)

  if color == '000000':
    color = ''

  return color

#SPECIFIC

def findLastDataRow(sheet):
  targetCol = sheet['first_col']

  lastRowOfDataFrame = len(targetCol) - 1
  
  return lastRowOfDataFrame

def getData(sheet, column, row):
    
  columns = sheet.columns

  value = ''

  if column in columns:
    value =  sheet[column][row]
  else:
    value =  ''

  # if type(value) is str and re.search('Caf', value) is not None:
  #   print(value)

  return str(value).replace("'", "''")
    
def getLenOfIterCols(wb, sheet):
  return len(list(wb[sheet].iter_cols()))

def getItemAtIndexOfIterCols(wb, sheet, i):
  from itertools import islice, count
  return next(islice(wb[sheet].iter_cols(), i, i+1))

def isfloat(value):
  try:
    float(value)
    return True
  except ValueError:
    return False

def getColumnNames():
  columns = []
  columns.append('first_col')
  columns.append('col_2')
  columns.append('col_3')
  columns.append('col_4')
  columns.append('col_5')
  # ...
  columns.append('last_col')

  return columns

#CHANGE THESE DEPENDING ON EXCEL DOC
nextGridSpecificId = 1
borderStartingPos = 14
tableName = 'example_table'
print('opening workbook...')
wb = op.load_workbook('example_workbook.XLSX')
print('opened')

print('get column names')
columnNames = getColumnNames()
print(columnNames)
df = collections.OrderedDict()
dfStyle = collections.OrderedDict()

print('start looping')
for sheetname in wb.sheetnames:

  print('sheet: ')
  print(sheetname)
  valueArr = collections.OrderedDict()
  valueArrStyle = collections.OrderedDict()

  for columnName in columnNames:
    valueArr[columnName] = []
    valueArrStyle[columnName] = []

  print('length of columnNames')
  print(len(columnNames))

  length = getLenOfIterCols(wb, sheetname)
  print('get length of iter cols')
  print(length)

  if length > 575:
    length = 575

  print('length changed')
  print(length)
    
  print('len(col)')
  print(len(getItemAtIndexOfIterCols(wb, sheetname, 0)))

  for i in range(0, length):
    col = getItemAtIndexOfIterCols(wb, sheetname, i)

    print('column ')
    print(i)

    for j in range(0, len(col)):
      cell = col[j]

      value = cell.value
      if value == None:
        value = ''

      value = str(value).replace(" 00:00:00", "")

      bgColor = getBgColor(wb, cell.fill)
      if bgColor != None and bgColor != '':
        bgColor = "#" + bgColor + "bg"
      else:
        bgColor = ''

      ftColor = getFtColor(wb, cell.font)
      if ftColor != None and ftColor != '':
        ftColor = "#" + ftColor + "ft"
      else:
        ftColor = ''

      bold = cell.font.b
      if bold != None and bold != False:
        bold = "$bld"
      else:
        bold = ''

      if i >= borderStartingPos:
        border = "$bdr"
      else:
        border = ''

      style = bgColor + ftColor + bold + border

      valueArr[columnNames[i]].append(value)
      valueArrStyle[columnNames[i]].append(style)

  df[sheetname] = pd.DataFrame(valueArr)
  dfStyle[sheetname] = pd.DataFrame(valueArrStyle)

  print('done')

#MAIN CODE START
firstDataColumn = 'first_col'
lastDataColumn = 'last_col'
dataStartPos = 0

query = ''

for sheetName in df:

  sheet = df[sheetName]
  sheetStyle = dfStyle[sheetName]

  dataEndPos = findLastDataRow(sheet)
  print('dataEndPos')
  print(dataEndPos)

  dataObj = collections.OrderedDict()
  styleObj = collections.OrderedDict()

  for columnName in columnNames:
    dataObj[columnName] = []
    styleObj[columnName] = []

  for row in range(dataStartPos, dataEndPos + 1):
    print('row#')
    print(row)
    
    #Get Data/styles
    for columnName in columnNames:
      dataObj[columnName].append(getData(sheet, columnName, row))
      styleObj[columnName].append(getData(sheetStyle, columnName, row))

    ndf = pd.DataFrame(dataObj)
    ndfStyles = pd.DataFrame(styleObj)

  print('got DataFrame, now do query')

  for row, tup in ndf[firstDataColumn].items():

    print('query row#')
    print(row)
    
    query += 'INSERT INTO ' + tableName + ' (grid_specific_id, '

    for column, tup in ndf.items():

      query += column

      if column != lastDataColumn:
        query += ','

    query += ') VALUES (' + str(nextGridSpecificId) + ', '

    for column, tup in ndf.items():

      query += "'" + str(ndf[column][row]) + "'"

      if column != lastDataColumn:
        query += ','

    query += ');\n'

    query += 'INSERT INTO ' + tableName + '_style (grid_specific_id, '

    for column, tup in ndf.items():

      query += 'styleattrib_' + column

      if column != lastDataColumn:
        query += ','

    query += ') VALUES (' + str(nextGridSpecificId) + ', '

    for column, tup in ndfStyles.items():

      query += "'" + str(ndfStyles[column][row]) + "'"

      if column != lastDataColumn:
        query += ','

    query += ');\n'

    nextGridSpecificId = nextGridSpecificId + 1

  print('done')

f = open("query.sql","w+")
f.write(query)
f.close